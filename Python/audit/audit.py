import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. データの読み込み
df_iam = pd.read_csv('~/Downloads/iam_list_detailed.csv')
df_sp = pd.read_csv('~/Downloads/Sharepoint_GoogleCloud権限一覧.csv')

# 2. Google Cloud側のデータ整形（接頭辞の剥離）
def extract_email(member):
    if ':' in str(member):
        return member.split(':', 1)[1]
    return member

def extract_type(member):
    if ':' in str(member):
        return member.split(':', 1)[0]
    return 'unknown'

df_iam['account_type'] = df_iam['member'].apply(extract_type)
df_iam['email_clean'] = df_iam['member'].apply(extract_email)

# deleted: などの特殊ケース（?uid=... のお尻）をカットして純粋なメールアドレスにする
df_iam['email_clean'] = df_iam['email_clean'].apply(lambda x: x.split('?')[0] if '?' in str(x) else x)

# 表記揺れを防ぐため小文字・スペース削除で統一
df_sp['mail_clean'] = df_sp['mail'].str.strip().str.lower()
df_iam['email_clean'] = df_iam['email_clean'].str.strip().str.lower()

# 3. 突合（マージ）
df_merged = pd.merge(df_iam, df_sp, left_on='email_clean', right_on='mail_clean', how='left')

# 4. 監査判定ロジックの定義
def judge_audit(row):
    ac_type = row['account_type']
    sp_mail = row['mail_clean']
    status = row['ApprovalStatus']
    role = row['role']
    
    if ac_type == 'serviceAccount':
        return '対象外 (システムアカウント)'
    elif ac_type == 'group':
        return '対象外 (Googleグループ)'
    elif ac_type == 'deleted':
        return '要削除 (削除済みアカウントの残骸)'
    
    # 人間（user:）の場合の判定
    if pd.isna(sp_mail):
        return '要確認 (SharePointに申請がないアカウント)'
    if status != '承認済み':
        return f'要確認 (申請ステータス: {status})'
    if any(p_role in str(role) for p_role in ['roles/owner', 'roles/editor', 'admin']):
        return '要レビュー (特権保有者)'
        
    return '問題なし (申請承認済み)'

df_merged['AuditResult'] = df_merged.apply(judge_audit, axis=1)

# 列を人間が見やすい順番に並び替え
output_cols = [
    'member', 'account_type', 'role', 'ApprovalStatus', 'Project', 'AccountName', 'StartDate', 'EndDate', 'AuditResult'
]
df_output = df_merged[output_cols].copy()
df_output.columns = [
    'Google Cloud メンバー', 'アカウント種別', '付与されているロール', 
    'SP申請ステータス', 'SP申請プロジェクト', 'SPアカウント名', '利用開始日', '利用終了日', '監査判定'
]

# 5. 見栄えの良いExcelワークブックの生成
wb = openpyxl.Workbook()
ws_sum = wb.active
ws_sum.title = "棚卸サマリー"
ws_sum.views.sheetView[0].showGridLines = True

# タイトル
ws_sum['A1'] = "Google Cloud IAM × SharePoint 突合棚卸報告書"
ws_sum['A1'].font = Font(name='Meiryo UI', size=16, bold=True, color='1B365D')
ws_sum['A3'] = "本シートは、エクスポートデータをプログラムによって自動名寄せ・判定したものです。"
ws_sum['A3'].font = Font(name='Meiryo UI', size=10, italic=True)

# 集計表の枠組み
ws_sum['A5'] = "監査結果の集計"
ws_sum['A5'].font = Font(name='Meiryo UI', size=12, bold=True, color='1B365D')

headers_sum = ["監査判定項目", "該当件数", "対応方針 / リスク"]
for col_num, header in enumerate(headers_sum, 1):
    cell = ws_sum.cell(row=6, column=col_num)
    cell.value = header
    cell.font = Font(name='Meiryo UI', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

summary_counts = df_output['監査判定'].value_counts()
all_statuses = [
    '要確認 (SharePointに申請がないアカウント)', '要レビュー (特権保有者)',
    '要削除 (削除済みアカウントの残骸)', '問題なし (申請承認済み)',
    '対象外 (システムアカウント)', '対象外 (Googleグループ)'
]

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

row_idx = 7
for status in all_statuses:
    count = summary_counts.get(status, 0)
    action = "問題ありません。"
    if "申請がない" in status:
        action = "不正アカウントまたは申請漏れの可能性があります。利用目的を確認し、未申請なら削除してください。"
    elif "特権保有者" in status:
        action = "開発環境であってもオーナー/編集者権限は最小限にすべきです。閲覧者等への格下げを検討してください。"
    elif "残骸" in status:
        action = "Google Cloud側で既に削除されたアカウントの権限定義だけが残っています。ポリシー清掃のため削除してください。"
    elif "システム" in status or "グループ" in status:
        action = "システム自動運用のためのアカウントです。通常の人事棚卸からは除外します。"
        
    ws_sum.cell(row=row_idx, column=1, value=status).border = thin_border
    ws_sum.cell(row=row_idx, column=2, value=count).border = thin_border
    ws_sum.cell(row=row_idx, column=3, value=action).border = thin_border
    ws_sum.cell(row=row_idx, column=2).alignment = Alignment(horizontal='right')
    
    if "要確認" in status or "要削除" in status:
        ws_sum.cell(row=row_idx, column=1).fill = PatternFill(start_color='FFD6D6', end_color='FFD6D6', fill_type='solid')
        ws_sum.cell(row=row_idx, column=1).font = Font(name='Meiryo UI', bold=True, color='9C0006')
    elif "要レビュー" in status:
        ws_sum.cell(row=row_idx, column=1).fill = PatternFill(start_color='FFF0C2', end_color='FFF0C2', fill_type='solid')
        ws_sum.cell(row=row_idx, column=1).font = Font(name='Meiryo UI', bold=True, color='9C6500')
    row_idx += 1

# 総計行
ws_sum.cell(row=row_idx, column=1, value="総アクセス権定義数").font = Font(name='Meiryo UI', bold=True)
ws_sum.cell(row=row_idx, column=1).border = thin_border
sum_cell = ws_sum.cell(row=row_idx, column=2, value=f"=SUM(B7:B{row_idx-1})")
sum_cell.font = Font(name='Meiryo UI', bold=True)
sum_cell.border = thin_border
sum_cell.alignment = Alignment(horizontal='right')
ws_sum.cell(row=row_idx, column=3).border = thin_border

# タブ2: 突合詳細
ws_det = wb.create_sheet(title="突合詳細データ")
ws_det.views.sheetView[0].showGridLines = True

for col_num, header in enumerate(df_output.columns, 1):
    cell = ws_det.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(name='Meiryo UI', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

for r_idx, row in enumerate(df_output.values, 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws_det.cell(row=r_idx, column=c_idx)
        cell.value = '' if pd.isna(val) else val
        cell.font = Font(name='Meiryo UI', size=10)
        cell.border = thin_border
        
        if c_idx == 9: # 監査判定列の着色
            val_str = str(val)
            if "要確認" in val_str or "要削除" in val_str:
                cell.fill = PatternFill(start_color='FFD6D6', end_color='FFD6D6', fill_type='solid')
                cell.font = Font(name='Meiryo UI', size=10, bold=True, color='9C0006')
            elif "要レビュー" in val_str:
                cell.fill = PatternFill(start_color='FFF0C2', end_color='FFF0C2', fill_type='solid')
                cell.font = Font(name='Meiryo UI', size=10, bold=True, color='9C6500')
            elif "問題なし" in val_str:
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

ws_det.auto_filter.ref = f"A1:I{len(df_output)+1}"

# 幅自動調整
for ws in [ws_sum, ws_det]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                byte_len = len(str(cell.value).encode('utf-8'))
                if byte_len > max_len: max_len = byte_len
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save("~/Downloads/GoogleCloud_SharePoint_突合棚卸結果.xlsx")
print("突合完了！「~/Downloads/GoogleCloud_SharePoint_突合棚卸結果.xlsx」を出力しました。")